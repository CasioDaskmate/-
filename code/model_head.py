import numpy as np
import matplotlib.pyplot as plt
import cv2 as cv
import os
from random import randint
import time
from openpyxl import Workbook
from scipy.optimize import curve_fit
import math

D = 756
number = 44

#二元高斯函数
def gaussian(xy, amplitude, center_x, center_y, sigma_x, sigma_y):
    x, y = xy[:, 0], xy[:, 1]
    return amplitude * np.exp(-((x - center_x) ** 2) / (2 * sigma_x ** 2) - ((y - center_y) ** 2) / (2 * sigma_y ** 2))

#背景建模
def background(image):
    frame = mask(image)
    #blur = cv.pyrMeanShiftFiltering(image,sp=60,sr=60)
    #d_frame = cv.cvtColor(blur,cv.COLOR_BGR2GRAY)
    #d_frame = cv.cvtColor(frame,cv.COLOR_BGR2GRAY)
    d_frame = cv.GaussianBlur(frame,(3,3),0)
    ret, binary = cv.threshold(d_frame, 0, 255, cv.THRESH_BINARY_INV)
    ret, binary = cv.threshold(binary, 0, 255, cv.THRESH_BINARY_INV+cv.THRESH_OTSU)
    kernel = cv.getStructuringElement(cv.MORPH_RECT, (1, 1))
    opening = cv.morphologyEx(binary, cv.MORPH_OPEN, kernel=kernel, iterations=1)
    sure_bg = cv.dilate(opening, kernel, iterations=1)
    cv.imshow('b',sure_bg)
    return sure_bg

#遮罩
def mask(frame):
    frame = cv.cvtColor(frame, cv.COLOR_BGR2GRAY)
    ret, frame = cv.threshold(frame, 0, 255, cv.THRESH_BINARY_INV+cv.THRESH_OTSU)
    height, width = frame.shape[:2]
    #print(height,width)
    center_x, center_y = int(D/2),int(D/2)
    #center_x, center_y = 371,371
    r = int(D/2 -30)
    mask = np.zeros((height, width), dtype=np.uint8)
    cv.circle(mask, (center_x, center_y), r, (255, 255, 255), -1)
    frame = cv.bitwise_and(frame, frame, mask=mask)
    cv.circle(mask, (center_x, center_y), r, (255, 255, 255), -1)
    ret, frame = cv.threshold(frame, 0, 255, cv.THRESH_BINARY_INV+cv.THRESH_OTSU)
    return frame

def cutsize(frame):
    #frame = frame[0:1010,685:1695]
    #frame = frame[10:1040,620:1650]
    #frame = frame[15:1075,590:1650]
    frame = frame[0:1080,450:1530] 
    frame = cv.resize(frame, None, fx=0.7, fy=0.7)
    return frame

#计算高斯中心，输出的是一个点
def gaussian_center(points):
    initial_params = [1.0, 300.0, 400.0, 100.0, 100.0]
    xy_data = np.array([points[:, 0], points[:, 1]]).T
    popt, _ = curve_fit(gaussian, xy_data, np.zeros(len(points)), p0=initial_params)
    gaussian_center_x, gaussian_center_y = popt[1], popt[2]
    return gaussian_center_x,gaussian_center_y


#在图中标出帧数
def frame_number(frame):
    cv.rectangle(frame, (10, 2), (100,20), (255,255,255), -1)
    cv.putText(frame, str(cap.get(cv.CAP_PROP_POS_FRAMES)), (15, 15),cv.FONT_HERSHEY_SIMPLEX, 0.5 , (0,0,0))
    return frame

def moving_state(v_in_ellipses_dict,x):
    v_list = list(v_in_ellipses_dict)
    extra_content = x
    modified_list = v_list + [extra_content]
    return modified_list

def turning_state(v_in_ellipses_dict,x):
    v_list = list(v_in_ellipses_dict)
    extra_content = x
    modified_list = v_list + [extra_content]
    return modified_list

#把数据变为以第n个为第一视角的数据
def pole_changing(ellipses_dict,n):
    pole_changing_dict = {}
    j = 0
    for i, v in ellipses_dict.items():
        if i != n:
            pole_changing_j = (float(((v[0][0] - ellipses_dict[n-1][0][0]) ** 2 + (v[0][1] - ellipses_dict[n-1][0][1]) ** 2) ** 0.5), abs(ellipses_dict[n-1][2] - angle_calculate(v[0][0],v[0][1],ellipses_dict[n-1][0][0],ellipses_dict[n-1][0][1])) ,ellipses_dict[i][3])
            pole_changing_dict[j] = pole_changing_j
            j = j + 1
    return pole_changing_dict

def angle_calculate(x1,y1,x2,y2):
    dx = x2 - x1
    dy = y2 - y1
    angle_rad = math.atan2(dy, dx)
    angle_deg = math.degrees(angle_rad)
    return angle_deg


cap = cv.VideoCapture("D:/科研/视频_2024.03.03/output_video_2.avi")
#cap = cv.VideoCapture("F:/科研/视频_2024.03.03/output_video_2.avi")
#cap = cv.VideoCapture("F:/科研/视频_2023.9.28/output_video_cut_1.avi")
#cap = cv.VideoCapture("F:/科研/视频_2023.10.13/output_video_cut_3.avi")
success, frame = cap.read()
#print(int(cap.get(cv.CAP_PROP_FRAME_WIDTH)))
#print(int(cap.get(cv.CAP_PROP_FRAME_HEIGHT)))
frame = cutsize(frame)
d_frame = background(frame)
counter_ids = []
colors = []
ellipses_dict = {}
contours = []
points = np.zeros((number, 2))
datas = np.zeros((int(cap.get(cv.CAP_PROP_FRAME_COUNT)),number))
last_datas = {}
all_the_datas = {}
row = 0
wb = Workbook()
ws = wb.create_sheet('平均距离变化表', 0)
dis_average = 0
count_for_head = 0
'''
#完成输出设置
time_t = time.strftime('%Y.%m.%d %H-%M-%S', time.localtime(time.time()))
outDir = 'F:/科研/平均距离分析_' + time_t
os.mkdir(outDir)
outFile = outDir + '平均距离变化_' + time_t + '.xlsx'
'''

while True:
    contours = []
    contours_tmp, hierarchy = cv.findContours(d_frame, cv.RETR_TREE,cv.CHAIN_APPROX_NONE)
    for i, c in enumerate(contours_tmp):
        if not (cv.arcLength(c,True) > 190 or cv.arcLength(c,True) < 20 or len(c) < 5):
            contours.append(c)

    for i, c in enumerate(contours):   
        if i not in counter_ids:
            x,y,w,h = cv.boundingRect(c)
            counter_ids.append(i)
            colors.append((randint(64, 255), randint(64, 255), randint(64, 255)))
            ellipse = cv.fitEllipse(c)
            ellipses_dict[i] = ellipse
            points[i][0] =  ellipses_dict[i][0][0]
            points[i][1] =  ellipses_dict[i][0][1]
            cv.ellipse(frame, ellipse, colors[i], 2)
            #cv.putText(frame, str(i+1), (int(ellipse[0][0]),int(ellipse[0][1])), cv.FONT_HERSHEY_PLAIN, 1, [0, 0, 255], 2)
    
    # 头尾识别
    for i, v in ellipses_dict.items():
        count_for_head = count_for_head + 1
        if count_for_head > number:
            break
        
        headn = 7
        pixel1 = frame[int(float(v[0][1]) + float(v[1][1]+ headn ) * 1/2 * math.sin(math.radians(float(v[2] - 90)))), int(float(v[0][0]) + float(v[1][1]+ headn ) * 1/2 * math.cos(math.radians(float(v[2] - 90))))]
        pixel2 = frame[int(float(v[0][1]) - float(v[1][1]+ headn ) * 1/2 * math.sin(math.radians(float(v[2] - 90)))), int(float(v[0][0]) - float(v[1][1]+ headn ) * 1/2 * math.cos(math.radians(float(v[2] - 90))))]
        gray_value1 = sum(pixel1) / 3
        gray_value2 = sum(pixel2) / 3
        # print(pixel1,pixel2)
        # print(gray_value1,gray_value2)
        if gray_value1 < gray_value2 :
            v_list = list(v)
            v_list[2] = v_list[2] + 180
            v = tuple(v_list)
            ellipses_dict[i] = tuple(v_list)
        cv.circle(frame,(int(float(v[0][0]) + float(v[1][1]) * 1/2 * math.cos(math.radians(float(v[2] - 90)))),int(float(v[0][1]) + float(v[1][1]) * 1/2 * math.sin(math.radians(float(v[2] - 90))))),3, (0, 0, 255), -1)



    gaussian_center_x,gaussian_center_y = gaussian_center(points)
    cv.circle(frame, (int(gaussian_center_x),int(gaussian_center_y)), 5, (0, 0, 255), -1)
    cv.imshow('first_step', frame)
    if cv.waitKey(1) == ord(' '):
        break

while cap.isOpened():   
    contour_ids_o = []
    success, frame = cap.read()
    frame = cutsize(frame)
    d_frame = background(frame)
    contours = []
    points = np.zeros((number, 2))
    contours_tmp, hierarchy = cv.findContours(d_frame, cv.RETR_TREE,cv.CHAIN_APPROX_NONE)
    frame = frame_number(frame)
    
    for i, c in enumerate(contours_tmp):
            if not (cv.arcLength(c,True) > 190 or cv.arcLength(c,True) < 20 or len(c) < 5):
                contours.append(c)

    if len(contours) == number:

        # 存入上一帧进last_datas，看情况是否要多存几帧
        for i,v in ellipses_dict.items():
            last_datas[i] = v

        all_the_datas[row] = last_datas

        for i, c in enumerate(contours):   
            x,y,w,h = cv.boundingRect(c)
            ellipse = cv.fitEllipse(c)
            nearest_ellipse = None
            min_distance = float('inf')
            for i, v in ellipses_dict.items():
                if i not in contour_ids_o:
                    distance = ((v[0][0] - ellipse[0][0]) ** 2 + (v[0][1] - ellipse[0][1]) ** 2) ** 0.5
                    if distance < min_distance:
                        min_distance = distance
                        nearest_ellipse = i

            if nearest_ellipse != None:      
                ellipses_dict[nearest_ellipse] = ellipse
                contour_ids_o.append(nearest_ellipse)
                cv.ellipse(frame, ellipse, colors[nearest_ellipse], 2)
                cv.putText(frame, str(nearest_ellipse+1), (int(ellipse[0][0]),int(ellipse[0][1])), cv.FONT_HERSHEY_PLAIN, 1, [0, 0, 255], 2)
                points[nearest_ellipse][0] =  ellipses_dict[nearest_ellipse][0][0]
                points[nearest_ellipse][1] =  ellipses_dict[nearest_ellipse][0][1]
        
        #判定运动情况，1表示运动，0表示静止,有两处可能需要修改：帧数差和dis_last_datas的标准
        for i,v in ellipses_dict.items():
            dis_last_datas = (((v[0][0] - last_datas[i][0][0]) ** 2 + (v[0][1] - last_datas[i][0][1]) ** 2) ** 0.5)
            if dis_last_datas > 0.1:
                ellipses_dict[i] = tuple(moving_state(v,1))
            else:
                ellipses_dict[i] = tuple(moving_state(v,0))

        for i,v in ellipses_dict.items():
            angle_change = v[2] - last_datas[i][2]
            if angle_change > 0.1 & angle_change != 180:
                ellipses_dict[i] = tuple(turning_state(v,1))
            else:
                ellipses_dict[i] = tuple(turning_state(v,0))

        gaussian_center_x,gaussian_center_y = gaussian_center(points)
        
        dis_total = 0
        dis_average = 0
        min_dis = float('inf')
        max_dis = 0
        for i,v in ellipses_dict.items():
            dis = (((v[0][0] - gaussian_center_x) ** 2 + (v[0][1] - gaussian_center_y) ** 2) ** 0.5)
            datas[row][i] = dis
            dis_total = dis_total + dis
            if dis > max_dis:
                max_dis = dis
            if dis < min_dis:
                min_dis = dis
        dis_total = dis_total - min_dis - max_dis   
        dis_average = dis_total/number-2

    # 头尾识别
    for i, v in ellipses_dict.items():        
        headn = 7
        pixel1 = frame[int(float(v[0][1]) + float(v[1][1]+ headn ) * 1/2 * math.sin(math.radians(float(v[2] - 90)))), int(float(v[0][0]) + float(v[1][1]+ headn ) * 1/2 * math.cos(math.radians(float(v[2] - 90))))]
        pixel2 = frame[int(float(v[0][1]) - float(v[1][1]+ headn ) * 1/2 * math.sin(math.radians(float(v[2] - 90)))), int(float(v[0][0]) - float(v[1][1]+ headn ) * 1/2 * math.cos(math.radians(float(v[2] - 90))))]
        gray_value1 = sum(pixel1) / 3
        gray_value2 = sum(pixel2) / 3
        # print(pixel1,pixel2)
        # print(gray_value1,gray_value2)
        if gray_value1 < gray_value2 :
            v_list = list(v)
            v_list[2] = v_list[2] + 180
            v = tuple(v_list)
            ellipses_dict[i] = tuple(v_list)
        cv.circle(frame,(int(float(v[0][0]) + float(v[1][1]) * 1/2 * math.cos(math.radians(float(v[2] - 90)))),int(float(v[0][1]) + float(v[1][1]) * 1/2 * math.sin(math.radians(float(v[2] - 90))))),3, (0, 0, 255), -1)




    ws.cell(row = row + 1, column = 1, value = str(cap.get(cv.CAP_PROP_POS_FRAMES)))
    ws.cell(row = row + 1, column = 2, value = str(dis_average)) 
    row = row + 1      
    cv.imshow('okframe', frame)
    for i,v in ellipses_dict.items():
        print(v[3])
    if cv.waitKey(1) == ord('q'):
        break

#print(datas)
#这里已经把所有数据输进datas了，直接用matlab画图就行了

print(all_the_datas)
#wb.save(str(outFile))   
cap.release()
cv.destroyAllWindows()